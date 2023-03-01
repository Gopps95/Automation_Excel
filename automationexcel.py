import openpyxl as xl
from openpyxl.chart import BarChart,Reference


wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
for row in range (1,sheet.max_row+1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 3
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price
Reference(sheet, min_row=2, max_row=sheet.max_row,min_col=4,max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('trance.xlsx')